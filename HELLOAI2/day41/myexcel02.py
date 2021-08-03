from openpyxl import load_workbook

 
load_wb = load_workbook("myexcel01.xlsx", data_only=True)
load_ws = load_wb['Sheet1']
# print(load_ws['B2'].value)


for i in range(2,2+3):
    print(load_ws.cell(i, 1).value ,end=" ")
    print(load_ws.cell(i, 2).value )
