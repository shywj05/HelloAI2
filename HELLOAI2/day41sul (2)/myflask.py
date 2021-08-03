from flask import Flask, render_template, jsonify, request, session, escape, redirect, send_file
import cx_Oracle
from day41sul.mydao_suser import MyDaoSuser
from day41sul.mydao_survey import MyDaoSurvey
from day41sul.mydao_sdetail import MyDaoSdetail
from day41sul.mydao_starget import MyDaoStarget
from day41sul.mydao_sresult import MyDaoSresult
from day41sul.mydao_snotice import MyDaoSnotice
from day41sul.mydao_sreply import MyDaoSreply


from day41sul.mymail import MyMail
from day41sul.mysms import MySms
from werkzeug.utils import secure_filename
import os
from openpyxl import load_workbook
from datetime import datetime

DIR_UPLOAD = "D:/workspace_python/HELLOAI2/day41sul/static/upload"

app = Flask(__name__, static_url_path="", static_folder="static")
app.secret_key = "ABCDEFG"



def getSession():
    user_id = ""
    try:
        user_id = str(escape(session["user_id"]))
    except:
        pass
    
    if user_id == "" :
        return False,user_id
    else :
        return True,user_id
    

@app.route("/")
@app.route("/main")
def main_render():
    return render_template('main.html')


@app.route("/sview")
def sview_render():
    survey_id = request.args.get('survey_id')
    st_mobile = request.args.get('st_mobile')
    list = MyDaoSdetail().myselect(survey_id)
    q_cnt= len(list)
    return render_template('sview.html', list=list ,survey_id=survey_id,st_mobile=st_mobile,q_cnt=q_cnt)

@app.route("/logout")
def logout_render():
    session.clear()
    return redirect("main")

#############################################

@app.route('/login.ajax', methods=['POST'])
def login_ajax():
    user_id = request.form["user_id"]
    pwd = request.form["pwd"]
    
    print("user_id",user_id)
    print("pwd",pwd)
    
    list = MyDaoSuser().mylogin(user_id, pwd)

    msg = ""
    if len(list) == 1:
        session["user_id"] = list[0]['user_id']
        session["user_name"] = list[0]['user_name']
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)
##################################################
@app.route('/join.ajax', methods=['POST'])
def join_ajax():
    user_id = request.form["user_id"]
    pwd = '1'
    user_name = request.form["user_name"]
    mobile = request.form["mobile"]
    email = request.form["email"]
    birth = request.form["birth"]
    
    print(user_id)
    
    cnt = MyDaoSuser().myinsert(user_id, pwd, user_name, mobile, email, birth, "", "S00001", "", "S00001")
    
    content = user_id+"님!!!^o^ 당신의 계정은 "+user_id+"이고 비밀번호는 1입니다. 로그인 후 변경바랍니다"
    MyMail().mysendmail(email, "설조 회원가입을 축하드립니다", content)

    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)



#############################################

@app.route('/dupl.ajax', methods=['POST'])
def dupl_ajax():
    user_id = request.form["user_id"]
    
    print("user_id",user_id)
    
    list = MyDaoSuser().mydupl(user_id)

    msg = ""
    if len(list) == 1:
        msg = "ng"
    else:
        msg = "ok"

    return jsonify(msg = msg)

#################################################

@app.route("/suser")
def suser_render():
    
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")

    
    list = MyDaoSuser().myselect()
    return render_template('suser.html', list=list)



@app.route('/suser_ins.ajax', methods=['POST'])
def suser_ins_ajax():
    user_id = request.form["user_id"]
    pwd = request.form["pwd"]
    user_name = request.form["user_name"]
    mobile = request.form["mobile"]
    email = request.form["email"]
    birth = request.form["birth"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))


    cnt = MyDaoSuser().myinsert(user_id, pwd, user_name, mobile, email, birth, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)



@app.route('/suser_upd.ajax', methods=['POST'])
def suser_upd_ajax():
    user_id = request.form["user_id"]
    pwd = request.form["pwd"]
    user_name = request.form["user_name"]
    mobile = request.form["mobile"]
    email = request.form["email"]
    birth = request.form["birth"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))
    cnt = MyDaoSuser().myupdate(user_id, pwd, user_name, mobile, email, birth, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)




@app.route('/suser_del.ajax', methods=['POST'])
def suser_del_ajax():
    user_id = request.form["user_id"]

    cnt = MyDaoSuser().mydelete(user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)



########################################################


@app.route("/survey")
def survey_render():
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    
    user_id = str(escape(session["user_id"]))
    
    list = MyDaoSurvey().myselect(user_id)
    return render_template('survey.html', list=list)


@app.route('/survey_ins.ajax', methods=['POST'])
def survey_ins_ajax():
    survey_id = request.form["survey_id"]
    title = request.form["title"]
    content = request.form["content"]
    interview_cnt = request.form["interview_cnt"]
    deadline = request.form["deadline"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))


    cnt = MyDaoSurvey().myinsert(survey_id, title, content, interview_cnt, deadline, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)



@app.route('/survey_upd.ajax', methods=['POST'])
def survey_upd_ajax():
    survey_id = request.form["survey_id"]
    title = request.form["title"]
    content = request.form["content"]
    interview_cnt = request.form["interview_cnt"]
    deadline = request.form["deadline"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))
    cnt = MyDaoSurvey().myupdate(survey_id, title, content, interview_cnt, deadline, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)


@app.route('/survey_del.ajax', methods=['POST'])
def survey_del_ajax():
    survey_id = request.form["survey_id"]

    cnt = MyDaoSurvey().mydelete(survey_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)



########################################################################


@app.route("/sdetail")
def sdetail_render():
    
    survey_id = request.args.get('survey_id')
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    
    list = MyDaoSdetail().myselect(survey_id)
    return render_template('sdetail.html', list=list ,survey_id=survey_id)


@app.route('/sdetail_ins.ajax', methods=['POST'])
def sdetail_ins_ajax():
    survey_id = request.form["survey_id"]
    s_seq = request.form["s_seq"]
    question = request.form["question"]
    a1 = request.form["a1"]
    a2 = request.form["a2"]
    a3 = request.form["a3"]
    a4 = request.form["a4"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))


    cnt = MyDaoSdetail().myinsert(survey_id, s_seq, question, a1, a2, a3, a4, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)


@app.route('/sdetail_upd.ajax', methods=['POST'])
def sdetail_upd_ajax():
    survey_id = request.form["survey_id"]
    s_seq = request.form["s_seq"]
    question = request.form["question"]
    a1 = request.form["a1"]
    a2 = request.form["a2"]
    a3 = request.form["a3"]
    a4 = request.form["a4"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))
    cnt = MyDaoSdetail().myupdate(survey_id, s_seq, question, a1, a2, a3, a4, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)



@app.route('/sdetail_del.ajax', methods=['POST'])
def sdetail_del_ajax():
    survey_id = request.form["survey_id"]
    s_seq = request.form["s_seq"]

    cnt = MyDaoSdetail().mydelete(survey_id, s_seq)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)

#####################################################################


@app.route("/starget")
def starget_render():
    survey_id = request.args.get('survey_id')
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    
    list = MyDaoStarget().myselect(survey_id)
    return render_template('starget.html', list=list,survey_id=survey_id)


@app.route("/starget_form")
def starget_form():
    survey_id = request.args.get('survey_id')
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    return render_template('starget_form.html',survey_id=survey_id)


@app.route("/starget_excel", methods=['POST'])
def starget_excel():
    survey_id = request.form["survey_id"]
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")  
    
    str_path_upload = "D:/workspace_python/HELLOAI2/day41sul/static/upload/"
    file = request.files['file']        
    filename = secure_filename(file.filename)
    file.save(str_path_upload+filename)
     
    load_wb = load_workbook(str_path_upload+filename, data_only=True)
    load_ws = load_wb['Sheet1']

    idx = 2
    cnt = 0
    while True:
        name_temp = load_ws.cell(idx, 1).value
        mobile_temp = load_ws.cell(idx, 2).value 
        if name_temp == None:
            break 
        print(name_temp,mobile_temp)
        mobile_temp = mobile_temp.replace("-","")
        cnt += MyDaoStarget().myinsert(survey_id, mobile_temp, "n", "", user_id, "", user_id)
        idx+=1
        
    return render_template('starget_excel.html',cnt=cnt)



@app.route('/starget_sms.ajax', methods=['POST'])
def starget_sms_ajax():
    survey_id = request.form["survey_id"]
    
    list = MyDaoStarget().myselect(survey_id)
    
    print(list)
    
    for i in list:
        mobile = i['st_mobile']
        url = "http://192.168.41.3:5000/sview?survey_id="+survey_id+"&st_mobile="+i['st_mobile']
        MySms().mysendSms(mobile, url)


    msg = ""
    if len(list) > 0:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)






@app.route('/starget_ins.ajax', methods=['POST'])
def starget_ins_ajax():
    survey_id = request.form["survey_id"]
    st_mobile = request.form["st_mobile"]
    complete_yn = request.form["complete_yn"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))


    cnt = MyDaoStarget().myinsert(survey_id, st_mobile, complete_yn, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)


@app.route('/starget_upd.ajax', methods=['POST'])
def starget_upd_ajax():
    survey_id = request.form["survey_id"]
    st_mobile = request.form["st_mobile"]
    complete_yn = request.form["complete_yn"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))
    cnt = MyDaoStarget().myupdate(survey_id, st_mobile, complete_yn, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)


@app.route('/starget_del.ajax', methods=['POST'])
def starget_del_ajax():
    survey_id = request.form["survey_id"]
    st_mobile = request.form["st_mobile"]

    cnt = MyDaoStarget().mydelete(survey_id, st_mobile)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)

#############################################################

@app.route("/sresult")
def sresult_render():
    survey_id = request.args.get('survey_id')
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    
    list = MyDaoSresult().myselect_graph(survey_id)
    return render_template('sresult.html', list=list,enumerate=enumerate)


@app.route('/sresult_submit.ajax', methods=['POST'])
def sresult_submit_ajax():
    survey_id = request.form["survey_id"]
    st_mobile = request.form["st_mobile"]
    ans = request.form["ans"]

    cnt = MyDaoSresult().myinsert_ans(survey_id, st_mobile, ans)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)



@app.route('/sresult_ins.ajax', methods=['POST'])
def sresult_ins_ajax():
    survey_id = request.form["survey_id"]
    s_seq = request.form["s_seq"]
    st_mobile = request.form["st_mobile"]
    answer = request.form["answer"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))


    cnt = MyDaoSresult().myinsert(survey_id, s_seq, st_mobile, answer, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)


@app.route('/sresult_upd.ajax', methods=['POST'])
def sresult_upd_ajax():
    survey_id = request.form["survey_id"]
    s_seq = request.form["s_seq"]
    st_mobile = request.form["st_mobile"]
    answer = request.form["answer"]
    in_date = request.form["in_date"]
    in_user_id = str(escape(session["user_id"]))
    up_date = request.form["up_date"]
    up_user_id = str(escape(session["user_id"]))
    cnt = MyDaoSresult().myupdate(survey_id, s_seq, st_mobile, answer, in_date, in_user_id, up_date, up_user_id)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)



@app.route('/sresult_del.ajax', methods=['POST'])
def sresult_del_ajax():
    survey_id = request.form["survey_id"]
    s_seq = request.form["s_seq"]
    st_mobile = request.form["st_mobile"]

    cnt = MyDaoSresult().mydelete(survey_id, s_seq, st_mobile)
    
    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)

#####################################################################


@app.route("/snotice_list")
def snotice_list_render():
    search = request.args.get('search','')
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    

    list = MyDaoSnotice().myselect_list(user_id,search)
    return render_template('snotice_list.html', list=list,enumerate=enumerate,search=search)


@app.route("/snotice_detail")
def snotice_detail_render():
    b_seq = request.args.get('b_seq')
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    

    daonotice = MyDaoSnotice()
    obj = daonotice.myselect(b_seq)
    daonotice.myhit(b_seq)
    return render_template('snotice_detail.html', notice=obj,enumerate=enumerate)


@app.route("/snotice_mod")
def snotice_mod_render():
    b_seq = request.args.get('b_seq')
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    

    obj = MyDaoSnotice().myselect(b_seq)
    return render_template('snotice_mod.html', notice=obj,enumerate=enumerate)

@app.route("/snotice_modact", methods=['POST'])
def snotice_modact_render():
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")

    b_seq           = request.form["b_seq"]
    display_yn      = request.form["display_yn"]
    title           = request.form["title"]
    content         = request.form["content"]
    attach_file_old = request.form["attach_file"]
    attach_path_old = request.form["attach_path"]
    
    if attach_file_old == 'None' :
        attach_file_old = ""
        attach_path_old = ""
    
    file = request.files['file']        
    attach_file_temp = secure_filename(file.filename)
    attach_path_temp = str(datetime.today().strftime("%Y%m%d%H%M%S")) +"_"+ attach_file_temp  
    file.save(os.path.join(DIR_UPLOAD, attach_path_temp))
    
    attach_path = ""
    attach_file = ""
    if file :
        print("__attach_path_temp",attach_path_temp)
        print("__attach_file_temp",attach_file_temp)
        attach_path = attach_path_temp 
        attach_file = attach_file_temp
        print("file O")
    else:
        attach_path = attach_path_old 
        attach_file = attach_file_old
        print("file X")
    
    cnt = MyDaoSnotice().myupdate(b_seq, display_yn, title, content, attach_path, attach_file, None, None, user_id, None, user_id)
    
#     obj = MyDaoSnotice().myupdate(b_seq, display_yn, title, content, attach_path, attach_file, None, None, user_id, None, user_id)
    return render_template('snotice_modact.html', cnt=cnt,b_seq=b_seq,enumerate=enumerate)


@app.route("/snotice_add")
def snotice_add_render():
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    return render_template('snotice_add.html',enumerate=enumerate)



@app.route("/snotice_addact", methods=['POST'])
def snotice_addact_render():

    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")   
    
    display_yn  = request.form["display_yn"]
    title       = request.form["title"]
    content     = request.form["content"]
    
    file = request.files['file']        
    attach_file_temp = secure_filename(file.filename)
    attach_path_temp = str(datetime.today().strftime("%Y%m%d%H%M%S")) +"_"+ attach_file_temp  
    file.save(os.path.join(DIR_UPLOAD, attach_path_temp))
    
    attach_path = ""
    attach_file = ""
    if file :
        attach_path = attach_path_temp 
        attach_file = attach_file_temp
        print("file O")
    else:
        print("file X")

    

    cnt = MyDaoSnotice().myinsert('', display_yn, title, content, attach_path, attach_file, '', '', user_id, '', user_id)

    return render_template('snotice_addact.html', cnt=cnt,enumerate=enumerate)



@app.route("/snotice_delact")
def snotice_delact_render():
    b_seq = request.args.get('b_seq')

    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    cnt = MyDaoSnotice().mydelete(b_seq)

    return render_template('snotice_delact.html', cnt=cnt,enumerate=enumerate)


@app.route('/snotice_del.ajax', methods=['POST'])
def snotice_del_ajax():
    b_seq = request.form["b_seq"]

    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    cnt = MyDaoSnotice().mydel_img(b_seq,user_id)

    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"
    return jsonify(msg = msg)



@app.route('/reply_add.ajax', methods=['POST'])
def reply_add_ajax():
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    
    b_seq = request.form["b_seq"]
    cmt = request.form["cmt"]
    
    cnt = MyDaoSreply().myinsert(b_seq, cmt, None, user_id, None, user_id)

    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)

@app.route('/reply_del.ajax', methods=['POST'])
def reply_del_ajax():
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    
    r_seq = request.form["r_seq"]
    
    cnt = MyDaoSreply().mydelete(r_seq)

    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)


@app.route('/reply_good.ajax', methods=['POST'])
def reply_good_ajax():
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    
    r_seq = request.form["r_seq"]
    
    cnt = MyDaoSreply().mygood(r_seq)

    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)

@app.route('/reply_bad.ajax', methods=['POST'])
def reply_bad_ajax():
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    
    r_seq = request.form["r_seq"]
    
    cnt = MyDaoSreply().mybad(r_seq)

    msg = ""
    if cnt == 1:
        msg = "ok"
    else:
        msg = "ng"

    return jsonify(msg = msg)




@app.route('/reply_list.ajax', methods=['POST'])
def reply_list_ajax():
    flag_ses, user_id = getSession()
    if not flag_ses:
        return redirect("login.html")    
    
    b_seq = request.form["b_seq"]
    
    list = MyDaoSreply().myselect(b_seq);

    return jsonify(list = list)



@app.route('/download')
def download():
    attach_path = request.args.get('attach_path')
    attach_file = request.args.get('attach_file')
    
    file_name = DIR_UPLOAD+"/"+attach_path
    return send_file(file_name,
                     mimetype='image/png',
                     attachment_filename=attach_file,
                     as_attachment=True)







#####################################################################

if __name__ == "__main__":
    app.run(host="0.0.0.0")


